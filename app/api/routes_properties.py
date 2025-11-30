"""Property management routes."""
from __future__ import annotations

import os
import shutil
import csv
import io
import math
from typing import List, Optional
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.dependencies import get_current_active_user, require_admin
from app.db.base import get_db
from app.models.property import Property, PropertyImage, PropertyStatus, PropertyType
from app.models.user import User, UserRole
from app.schemas.property import (
    PropertyCreate,
    PropertyResponse,
    PropertyUpdate,
    PropertyComparisonResponse,
    PropertyComparisonItem,
    PropertyComparisonMetrics,
    PropertyComparisonSummary,
    PropertyMapResponse,
    PropertyMapPoint,
    PropertyImportResult,
)

router = APIRouter(prefix="/api/v1/properties", tags=["properties"])

# Configure upload directory
UPLOAD_DIR = Path("static/uploads/properties")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


def _haversine_distance(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calculate the great-circle distance between two points on Earth in miles."""
    radius_earth_miles = 3958.8
    lat1_rad, lon1_rad, lat2_rad, lon2_rad = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2_rad - lat1_rad
    dlon = lon2_rad - lon1_rad
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1_rad) * math.cos(lat2_rad) * math.sin(dlon / 2) ** 2
    c = 2 * math.asin(math.sqrt(a))
    return radius_earth_miles * c


def _filter_properties(
    *,
    query,
    current_user: User,
    owner_user_id: Optional[int] = None,
    city: Optional[str] = None,
    state: Optional[str] = None,
    zip_code: Optional[str] = None,
    property_type: Optional[PropertyType] = None,
    status: Optional[List[PropertyStatus]] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    min_bedrooms: Optional[int] = None,
    max_bedrooms: Optional[int] = None,
    min_bathrooms: Optional[float] = None,
    max_bathrooms: Optional[float] = None,
    min_square_feet: Optional[int] = None,
    max_square_feet: Optional[int] = None,
    min_year_built: Optional[int] = None,
    max_year_built: Optional[int] = None,
    tags: Optional[List[str]] = None,
    latitude: Optional[float] = None,
    longitude: Optional[float] = None,
    radius_miles: Optional[float] = None,
) -> List[Property]:
    """Apply filtering options to a property query and return matching records."""

    if current_user.role != UserRole.ADMIN:
        query = query.filter(Property.owner_user_id == current_user.id)
    elif owner_user_id:
        query = query.filter(Property.owner_user_id == owner_user_id)

    if city:
        query = query.filter(Property.city.ilike(f"%{city}%"))
    if state:
        query = query.filter(Property.state == state)
    if zip_code:
        query = query.filter(Property.zip_code == zip_code)
    if property_type:
        query = query.filter(Property.property_type == property_type)
    if status:
        query = query.filter(Property.status.in_(status))
    if min_price is not None:
        query = query.filter(Property.list_price >= min_price)
    if max_price is not None:
        query = query.filter(Property.list_price <= max_price)
    if min_bedrooms is not None:
        query = query.filter(Property.bedrooms >= min_bedrooms)
    if max_bedrooms is not None:
        query = query.filter(Property.bedrooms <= max_bedrooms)
    if min_bathrooms is not None:
        query = query.filter(Property.bathrooms >= min_bathrooms)
    if max_bathrooms is not None:
        query = query.filter(Property.bathrooms <= max_bathrooms)
    if min_square_feet is not None:
        query = query.filter(Property.square_feet >= min_square_feet)
    if max_square_feet is not None:
        query = query.filter(Property.square_feet <= max_square_feet)
    if min_year_built is not None:
        query = query.filter(Property.year_built >= min_year_built)
    if max_year_built is not None:
        query = query.filter(Property.year_built <= max_year_built)

    filtered = query.all()

    if tags:
        tag_set = {t.lower() for t in tags}
        filtered = [
            prop for prop in filtered
            if prop.tags and tag_set.issubset({t.lower() for t in prop.tags})
        ]

    if radius_miles and latitude is not None and longitude is not None:
        filtered = [
            prop for prop in filtered
            if prop.latitude is not None
            and prop.longitude is not None
            and _haversine_distance(latitude, longitude, prop.latitude, prop.longitude) <= radius_miles
        ]

    return filtered

@router.post("", response_model=PropertyResponse, status_code=status.HTTP_201_CREATED)
def create_property(
    property_data: PropertyCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> PropertyResponse:
    """Create a new property."""
    db_property = Property(
        owner_user_id=current_user.id,
        **property_data.model_dump(),
    )
    db.add(db_property)
    db.commit()
    db.refresh(db_property)
    return PropertyResponse.model_validate(db_property)


@router.get("", response_model=List[PropertyResponse])
def list_properties(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=100),
    city: Optional[str] = None,
    state: Optional[str] = Query(None, min_length=2, max_length=2),
    zip_code: Optional[str] = None,
    property_type: Optional[PropertyType] = None,
    status: Optional[List[PropertyStatus]] = Query(None),
    min_price: Optional[float] = Query(None, ge=0),
    max_price: Optional[float] = Query(None, ge=0),
    min_bedrooms: Optional[int] = Query(None, ge=0),
    max_bedrooms: Optional[int] = Query(None, ge=0),
    min_bathrooms: Optional[float] = Query(None, ge=0),
    max_bathrooms: Optional[float] = Query(None, ge=0),
    min_square_feet: Optional[int] = Query(None, ge=0),
    max_square_feet: Optional[int] = Query(None, ge=0),
    min_year_built: Optional[int] = Query(None, ge=1800),
    max_year_built: Optional[int] = Query(None, le=2100),
    owner_user_id: Optional[int] = None,
    tags: Optional[List[str]] = Query(None),
    latitude: Optional[float] = Query(None, ge=-90, le=90),
    longitude: Optional[float] = Query(None, ge=-180, le=180),
    radius_miles: Optional[float] = Query(None, gt=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> List[PropertyResponse]:
    """List properties for current user (or all if admin) with advanced filtering."""
    query = db.query(Property)
    filtered = _filter_properties(
        query=query,
        current_user=current_user,
        owner_user_id=owner_user_id,
        city=city,
        state=state,
        zip_code=zip_code,
        property_type=property_type,
        status=status,
        min_price=min_price,
        max_price=max_price,
        min_bedrooms=min_bedrooms,
        max_bedrooms=max_bedrooms,
        min_bathrooms=min_bathrooms,
        max_bathrooms=max_bathrooms,
        min_square_feet=min_square_feet,
        max_square_feet=max_square_feet,
        min_year_built=min_year_built,
        max_year_built=max_year_built,
        tags=tags,
        latitude=latitude,
        longitude=longitude,
        radius_miles=radius_miles,
    )

    paginated = filtered[skip: skip + limit]
    return [PropertyResponse.model_validate(prop) for prop in paginated]


@router.get("/compare", response_model=PropertyComparisonResponse)
def compare_properties(
    property_ids: List[int] = Query(..., min_items=2, max_items=5),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> PropertyComparisonResponse:
    """Compare up to 5 properties side-by-side with key metrics."""

    query = db.query(Property).filter(Property.id.in_(property_ids))
    if current_user.role != UserRole.ADMIN:
        query = query.filter(Property.owner_user_id == current_user.id)

    properties = query.all()
    if len(properties) != len(set(property_ids)):
        raise HTTPException(status_code=404, detail="One or more properties not found")

    ordered_props = sorted(properties, key=lambda p: property_ids.index(p.id))
    comparison_items: List[PropertyComparisonItem] = []

    price_values = []
    price_per_sqft_values = []
    sqft_values = []
    bedroom_values = []
    bathroom_values = []
    year_values = []
    tag_set = set()
    statuses = set()

    for prop in ordered_props:
        price_per_sqft = None
        if prop.list_price and prop.square_feet:
            price_per_sqft = round(prop.list_price / prop.square_feet, 2)
            price_per_sqft_values.append(price_per_sqft)

        metrics = PropertyComparisonMetrics(
            price_per_sqft=price_per_sqft,
            bedroom_count=prop.bedrooms,
            bathroom_count=prop.bathrooms,
            square_feet=prop.square_feet,
            list_price=prop.list_price,
            year_built=prop.year_built,
        )

        base = PropertyResponse.model_validate(prop)
        comparison_items.append(
            PropertyComparisonItem(
                **base.model_dump(),
                comparison_metrics=metrics,
            )
        )

        if prop.list_price is not None:
            price_values.append(prop.list_price)
        sqft_values.append(prop.square_feet)
        bedroom_values.append(prop.bedrooms)
        bathroom_values.append(prop.bathrooms)
        if prop.year_built:
            year_values.append(prop.year_built)
        if prop.tags:
            tag_set.update(prop.tags)
        statuses.add(prop.status)

    summary = PropertyComparisonSummary(
        average_price=round(sum(price_values) / len(price_values), 2) if price_values else None,
        average_price_per_sqft=round(sum(price_per_sqft_values) / len(price_per_sqft_values), 2)
        if price_per_sqft_values
        else None,
        average_square_feet=round(sum(sqft_values) / len(sqft_values), 2) if sqft_values else None,
        bedroom_range=f"{min(bedroom_values)}-{max(bedroom_values)}" if bedroom_values else None,
        bathroom_range=f"{min(bathroom_values)}-{max(bathroom_values)}" if bathroom_values else None,
        year_built_range=f"{min(year_values)}-{max(year_values)}" if year_values else None,
        tags=sorted(tag_set),
        statuses=sorted(statuses, key=lambda s: s.value),
    )

    return PropertyComparisonResponse(properties=comparison_items, summary=summary)


@router.get("/map", response_model=PropertyMapResponse)
def map_properties(
    city: Optional[str] = None,
    state: Optional[str] = Query(None, min_length=2, max_length=2),
    zip_code: Optional[str] = None,
    property_type: Optional[PropertyType] = None,
    status: Optional[List[PropertyStatus]] = Query(None),
    owner_user_id: Optional[int] = None,
    tags: Optional[List[str]] = Query(None),
    latitude: Optional[float] = Query(None, ge=-90, le=90),
    longitude: Optional[float] = Query(None, ge=-180, le=180),
    radius_miles: Optional[float] = Query(None, gt=0),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> PropertyMapResponse:
    """Interactive map-friendly representation of properties."""

    filtered = _filter_properties(
        query=db.query(Property),
        current_user=current_user,
        owner_user_id=owner_user_id,
        city=city,
        state=state,
        zip_code=zip_code,
        property_type=property_type,
        status=status,
        tags=tags,
        latitude=latitude,
        longitude=longitude,
        radius_miles=radius_miles,
    )

    points = [
        PropertyMapPoint(
            id=prop.id,
            latitude=prop.latitude,
            longitude=prop.longitude,
            label=f"{prop.address_line1}, {prop.city}, {prop.state}",
            property_type=prop.property_type,
            status=prop.status,
            heat_metric="list_price" if prop.list_price is not None else "square_feet",
            heat_value=prop.list_price if prop.list_price is not None else float(prop.square_feet),
            key_metrics={
                "bedrooms": prop.bedrooms,
                "bathrooms": prop.bathrooms,
                "square_feet": prop.square_feet,
                "year_built": prop.year_built,
            },
        )
        for prop in filtered
    ]

    return PropertyMapResponse(points=points)


@router.get("/{property_id}", response_model=PropertyResponse)
def get_property(
    property_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> PropertyResponse:
    """Get a property by ID."""
    property_obj = db.query(Property).filter(Property.id == property_id).first()
    if not property_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Property not found",
        )

    # Check ownership (unless admin)
    if current_user.role != UserRole.ADMIN and property_obj.owner_user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions",
        )

    return PropertyResponse.model_validate(property_obj)


@router.put("/{property_id}", response_model=PropertyResponse)
def update_property(
    property_id: int,
    property_data: PropertyUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> PropertyResponse:
    """Update a property."""
    property_obj = db.query(Property).filter(Property.id == property_id).first()
    if not property_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Property not found",
        )

    # Check ownership (unless admin)
    if current_user.role != UserRole.ADMIN and property_obj.owner_user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions",
        )

    # Update fields
    update_data = property_data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(property_obj, field, value)

    db.commit()
    db.refresh(property_obj)
    return PropertyResponse.model_validate(property_obj)


@router.delete("/{property_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_property(
    property_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> None:
    """Delete a property."""
    property_obj = db.query(Property).filter(Property.id == property_id).first()
    if not property_obj:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Property not found",
        )

    # Check ownership (unless admin)
    if current_user.role != UserRole.ADMIN and property_obj.owner_user_id != current_user.id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions",
        )

    db.delete(property_obj)
    db.commit()


@router.post("/import", response_model=PropertyImportResult)
def import_properties(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> PropertyImportResult:
    """Bulk import properties via CSV or Excel."""

    filename = file.filename or ""
    content = file.file.read()

    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        rows = [dict(zip(headers, [cell.value for cell in row])) for row in sheet.iter_rows(min_row=2)]
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV or Excel.")

    required_fields = {"address_line1", "city", "state", "zip_code", "property_type", "bedrooms", "bathrooms", "square_feet"}
    created_ids: List[int] = []
    errors: List[str] = []

    for idx, row in enumerate(rows, start=1):
        missing = required_fields - {k for k, v in row.items() if v not in (None, "")}
        if missing:
            errors.append(f"Row {idx}: missing fields {sorted(missing)}")
            continue

        try:
            prop = Property(
                owner_user_id=current_user.id,
                address_line1=str(row.get("address_line1")),
                address_line2=row.get("address_line2"),
                city=str(row.get("city")),
                state=str(row.get("state")),
                zip_code=str(row.get("zip_code")),
                country=row.get("country") or "USA",
                property_type=PropertyType(str(row.get("property_type")).lower()),
                status=PropertyStatus(str(row.get("status", PropertyStatus.EVALUATING.value)).lower()),
                bedrooms=int(row.get("bedrooms")),
                bathrooms=float(row.get("bathrooms")),
                square_feet=int(row.get("square_feet")),
                year_built=int(row.get("year_built")) if row.get("year_built") else None,
                list_price=float(row.get("list_price")) if row.get("list_price") else None,
                tags=row.get("tags").split("|") if row.get("tags") else None,
                latitude=float(row.get("latitude")) if row.get("latitude") else None,
                longitude=float(row.get("longitude")) if row.get("longitude") else None,
            )
        except Exception as exc:  # noqa: BLE001
            errors.append(f"Row {idx}: {exc}")
            continue

        db.add(prop)
        db.flush()
        created_ids.append(prop.id)

    db.commit()

    return PropertyImportResult(
        imported=len(created_ids),
        skipped=len(rows) - len(created_ids),
        errors=errors,
        created_ids=created_ids,
    )


@router.get("/export")
def export_properties(
    format: str = Query("csv", pattern="^(csv|excel)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    city: Optional[str] = None,
    state: Optional[str] = Query(None, min_length=2, max_length=2),
    zip_code: Optional[str] = None,
    property_type: Optional[PropertyType] = None,
    status: Optional[List[PropertyStatus]] = Query(None),
    owner_user_id: Optional[int] = None,
) -> StreamingResponse:
    """Export filtered properties to CSV or Excel for reporting."""

    filtered = _filter_properties(
        query=db.query(Property),
        current_user=current_user,
        owner_user_id=owner_user_id,
        city=city,
        state=state,
        zip_code=zip_code,
        property_type=property_type,
        status=status,
    )

    fieldnames = [
        "id",
        "address_line1",
        "address_line2",
        "city",
        "state",
        "zip_code",
        "country",
        "property_type",
        "status",
        "bedrooms",
        "bathrooms",
        "square_feet",
        "year_built",
        "list_price",
        "tags",
        "latitude",
        "longitude",
    ]

    if format == "csv":
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=fieldnames)
        writer.writeheader()
        for prop in filtered:
            writer.writerow({
                "id": prop.id,
                "address_line1": prop.address_line1,
                "address_line2": prop.address_line2,
                "city": prop.city,
                "state": prop.state,
                "zip_code": prop.zip_code,
                "country": prop.country,
                "property_type": prop.property_type.value,
                "status": prop.status.value,
                "bedrooms": prop.bedrooms,
                "bathrooms": prop.bathrooms,
                "square_feet": prop.square_feet,
                "year_built": prop.year_built,
                "list_price": prop.list_price,
                "tags": "|".join(prop.tags or []),
                "latitude": prop.latitude,
                "longitude": prop.longitude,
            })
        buffer.seek(0)
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=properties.csv"},
        )

    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(fieldnames)
    for prop in filtered:
        sheet.append([
            prop.id,
            prop.address_line1,
            prop.address_line2,
            prop.city,
            prop.state,
            prop.zip_code,
            prop.country,
            prop.property_type.value,
            prop.status.value,
            prop.bedrooms,
            prop.bathrooms,
            prop.square_feet,
            prop.year_built,
            prop.list_price,
            "|".join(prop.tags or []),
            prop.latitude,
            prop.longitude,
        ])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=properties.xlsx"},
    )


@router.post("/{property_id}/images", status_code=status.HTTP_201_CREATED)
def upload_property_image(
    property_id: int,
    file: UploadFile = File(...),
    is_primary: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Upload an image for a property."""
    property_obj = db.query(Property).filter(Property.id == property_id).first()
    if not property_obj:
        raise HTTPException(status_code=404, detail="Property not found")
        
    # Check ownership
    if current_user.role != UserRole.ADMIN and property_obj.owner_user_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    # Validate file type
    if not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="File must be an image")

    # Generate filename
    file_ext = os.path.splitext(file.filename)[1]
    filename = f"prop_{property_id}_{os.urandom(4).hex()}{file_ext}"
    file_path = UPLOAD_DIR / filename
    
    # Save file
    try:
        with file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not save file: {e}")

    # Create URL (assuming served from /static)
    image_url = f"/static/uploads/properties/{filename}"
    
    # Save to DB
    image = PropertyImage(
        property_id=property_id,
        url=image_url,
        is_primary=is_primary
    )
    db.add(image)
    db.commit()
    
    return {"url": image_url, "id": image.id}
